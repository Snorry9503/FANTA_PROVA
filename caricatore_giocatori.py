import openpyxl
import sqlite3
import datetime

#Creazione di classe Calciatore

class Calciatore:
   
   def __init__(self,
                anno,
                id,
                ruolo,
                ruoloMantra,
                nome,
                squadra,
                squadraFanta,
                pv,
                mv,
                fm,
                gf,
                gs,
                rp,
                rc,
                rplus,
                rminus,
                assenze,
                amm,
                esp,
                au):
    self.anno=anno    
    self.id=id
    self.ruolo = ruolo
    self.ruoloMantra = ruoloMantra
    self.nome = nome
    self.squadra = squadra
    self.squadraFanta = squadraFanta
    self.pv = pv 
    self.mv = mv 
    self.fm = fm 
    self.gf = gf 
    self.gs = gs 
    self.rp = rp 
    self.rc = rc 
    self.rplus   = rplus  
    self.rminus  = rminus 
    self.assenze = assenze
    self.amm     = amm    
    self.esp     = esp    
    self.au      = au 



def caricaGiocatoreAP(directory_path):

   dataAnnoCalcistico=datetime.datetime.now()  
#Caricamento directory Padre Excel

   statAP = openpyxl.load_workbook(f"{directory_path}/StatisticheAP.xlsx")

#Caricamento file foglio Excel

   sheetStatAP = statAP.active
    
#Array di classi Calciatore
   giocatoriAp=[]

#Loop di scorrimento su colonne

   for rowAp in sheetStatAP.iter_rows(min_row=3, values_only=True):
    giocatoriAp.append(Calciatore(
                            dataAnnoCalcistico.year-1,
                            rowAp[0],
                            rowAp[1],
                            rowAp[2],
                            rowAp[3],
                            rowAp[4],
                            None,
                            rowAp[5],
                            rowAp[6],
                            rowAp[7],
                            rowAp[8],
                            rowAp[9],
                            rowAp[10],
                            rowAp[11],
                            rowAp[12],
                            rowAp[13],
                            rowAp[14],
                            rowAp[15],
                            rowAp[16],
                            rowAp[17],
                            ))
    return giocatoriAp
   
def caricaGiocatoreAP(directory_path):
   dataAnnoCalcistico=datetime.datetime.now()
   #Caricamento directory Padre Excel

   statAC = openpyxl.load_workbook(f"{directory_path}/StatisticheAC.xlsx")

#Caricamento file foglio Excel

   sheetStatAC = statAC.active
    
#Array di classi Calciatore

   giocatoriAc=[]

#Loop di caricamento

   for rowAC in sheetStatAC.iter_rows(min_row=3, values_only=True):
            giocatoriAc.append(Calciatore(
                                dataAnnoCalcistico.year,
                                rowAC[0],
                                rowAC[1],
                                rowAC[2],
                                rowAC[3],
                                rowAC[4],
                                None,
                                rowAC[5],
                                rowAC[6],
                                rowAC[7],
                                rowAC[8],
                                rowAC[9],
                                rowAC[10],
                                rowAC[11],
                                rowAC[12],
                                rowAC[13],
                                rowAC[14],
                                rowAC[15],
                                rowAC[16],
                                rowAC[17],
                                ))
            return giocatoriAc
        
def crea_database(nome_db):
    conn = sqlite3.connect(nome_db)
    cursor = conn.cursor()
    
    # Creazione della tabella dei giocatori anno corrente
    cursor.execute('''CREATE TABLE IF NOT EXISTS giocatoriac (
	id INTEGER PRIMARY KEY ,
    ruolo TEXT	    		,
    ruoloMantra	 TEXT		,
    nome		 TEXT		,
    squadra		 TEXT		,
    squadraFanta TEXT		,
    pv			 INTEGER		,
    mv			 INTEGER		,
    fm			 INTEGER		,
    gf			 INTEGER		,
    gs			 INTEGER		,
    rp			 INTEGER		,
    rc			 INTEGER		,
    rplus		 INTEGER		,
    rminus		 INTEGER		,
    assenze		 INTEGER		,
    amm			 INTEGER		,
    esp			 INTEGER		,
    au           INTEGER)
	 ''')
    
    # Creazione della tabella dei giocatori anno precedente
    cursor.execute('''CREATE TABLE IF NOT EXISTS giocatoriap (
	id INTEGER PRIMARY KEY ,
    ruolo TEXT			,
    ruoloMantra	 TEXT		,
    nome		 TEXT		,
    squadra		 TEXT		,
    squadraFanta TEXT		,
    pv			 INTEGER		,
    mv			 INTEGER		,
    fm			 INTEGER		,
    gf			 INTEGER		,
    gs			 INTEGER		,
    rp			 INTEGER		,
    rc			 INTEGER		,
    rplus		 INTEGER		,
    rminus		 INTEGER		,
    assenze		 INTEGER		,
    amm			 INTEGER		,
    esp			 INTEGER		,
    au           INTEGER)
	 ''')
    conn.commit()
    conn.close()

def popolaDatabase(
                nomedb, 
                nometabella,
                id,
                ruolo,
                ruoloMantra,
                nome,
                squadra,
                squadraFanta,
                pv,
                mv,
                fm,
                gf,
                gs,
                rp,
                rc,
                rplus,
                rminus,
                assenze,
                amm,
                esp,
                au):
   conn = sqlite3.connect(nomedb)
   cursor = conn.cursor()
   if nometabella == 'giocatoriac':
      cursor.execute('''
        INSERT INTO giocatoriac ( 
                id,
                ruolo,
                ruoloMantra,
                nome,
                squadra,
                squadraFanta,
                pv,
                mv,
                fm,
                gf,
                gs,
                rp,
                rc,
                rplus,
                rminus,
                assenze,
                amm,
                esp,
                au)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
                id,
                ruolo,
                ruoloMantra,
                nome,
                squadra,
                squadraFanta,
                pv,
                mv,
                fm,
                gf,
                gs,
                rp,
                rc,
                rplus,
                rminus,
                assenze,
                amm,
                esp,
                au))
   
   
      conn.commit()
      conn.close()
      
   elif nometabella == 'giocatoriap':
      cursor.execute('''
        INSERT INTO giocatoriap ( 
                id,
                ruolo,
                ruoloMantra,
                nome,
                squadra,
                squadraFanta,
                pv,
                mv,
                fm,
                gf,
                gs,
                rp,
                rc,
                rplus,
                rminus,
                assenze,
                amm,
                esp,
                au)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
                id,
                ruolo,
                ruoloMantra,
                nome,
                squadra,
                squadraFanta,
                pv,
                mv,
                fm,
                gf,
                gs,
                rp,
                rc,
                rplus,
                rminus,
                assenze,
                amm,
                esp,
                au))
   
   conn.commit()
   conn.close()   

def cancellaDatabase(nomedb):
   conn = sqlite3.connect(nomedb)
   cursor = conn.cursor()

   cursor.execute(''' TRUNCATE TABLE giocatoriap''')
   cursor.execute(''' TRUNCATE TABLE giocatoriac''')

   conn.commit()
   conn.close()